from xml.etree import ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

with open("arquivo.xml", "rt", encoding="utf8") as f:
   tree = ET.parse(f)
   filename = f.name
root = tree.getroot()

ns = { 'cumulus' : 'http://www.canto.com/ns/Export/1.0'}

wb = Workbook()
ws = wb.active
ws.title = 'Cumulus'

def mk_header():
   """preenche primeira linha da planilha

      com os campos exportados, retornando
      dicionário com uid : nome do campo"""
   row = 1
   col = 'A'
   uidcampo = {}
   for node in root[0][0].findall('cumulus:Field', ns):
      #print(node.attrib, node.text)
      campo = node.find('cumulus:Name', ns)
      ws[ col + str(row) ] = campo.text
      col = get_column_letter(column_index_from_string(col) + 1)
      uidcampo[node.attrib['uid']] = campo.text
   return uidcampo

def fill_record(dic):
   row = 2
   count = 0
   for ITEM in root[1]:
      
      for FieldValue in ITEM.findall('cumulus:FieldValue', ns):
         if FieldValue.attrib['uid'] in dic.keys():
            chave = FieldValue.attrib['uid']
            for campo in FieldValue.iter():
                col = get_column_letter(list(dic.keys()).index(chave) + 1)
                ws[ col + str(row)] = campo.text
      count += 1
      row += 1
   return count


dic = mk_header()
itens = fill_record(dic)
print('Você converteu', len(dic), 'campos e preencheu', itens, 'itens.')

wb.save(str(filename)[:-4]+'.xlsx')
